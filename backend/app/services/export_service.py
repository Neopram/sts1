"""
Advanced Export Service - Phase 2
Supports multiple export formats: JSON, CSV, XML, PDF, Excel, SQL
"""

import json
import csv
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
from io import StringIO, BytesIO
from datetime import datetime
from typing import Dict, List, Any, Tuple
import logging

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportService:
    """Handle exports in multiple formats"""
    
    SUPPORTED_FORMATS = ["json", "csv", "xml", "pdf", "xlsx", "sql"]
    
    @staticmethod
    def export_to_json(data: Dict) -> Tuple[str, str]:
        """
        Export data as JSON
        
        Returns:
            Tuple of (content, filename)
        """
        try:
            content = json.dumps(data, indent=2, default=str)
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            return content, filename
        except Exception as e:
            logger.error(f"Error exporting to JSON: {str(e)}")
            raise
    
    @staticmethod
    def export_to_csv(data: Dict, headers: List[str] = None) -> Tuple[str, str]:
        """
        Export data as CSV
        
        Args:
            data: Data dictionary
            headers: Column headers
            
        Returns:
            Tuple of (content, filename)
        """
        try:
            output = StringIO()
            
            # Flatten data structure for CSV
            if isinstance(data, dict):
                # Single record
                if headers is None:
                    headers = list(data.keys())
                writer = csv.DictWriter(output, fieldnames=headers)
                writer.writeheader()
                writer.writerow({k: data.get(k, "") for k in headers})
            elif isinstance(data, list):
                # Multiple records
                if not data:
                    return "", "export.csv"
                
                if headers is None:
                    headers = list(data[0].keys())
                
                writer = csv.DictWriter(output, fieldnames=headers)
                writer.writeheader()
                for row in data:
                    writer.writerow({k: row.get(k, "") for k in headers})
            
            content = output.getvalue()
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            return content, filename
        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            raise
    
    @staticmethod
    def export_to_xml(data: Dict, root_name: str = "root") -> Tuple[str, str]:
        """
        Export data as XML
        
        Args:
            data: Data dictionary
            root_name: Root element name
            
        Returns:
            Tuple of (content, filename)
        """
        try:
            def dict_to_xml(data: Any, parent: ET.Element):
                """Convert dict to XML elements"""
                if isinstance(data, dict):
                    for key, value in data.items():
                        # Make key XML-safe
                        safe_key = key.replace("-", "_").replace(" ", "_")
                        child = ET.SubElement(parent, safe_key)
                        dict_to_xml(value, child)
                elif isinstance(data, list):
                    for item in data:
                        child = ET.SubElement(parent, "item")
                        dict_to_xml(item, child)
                else:
                    parent.text = str(data)
            
            root = ET.Element(root_name)
            root.set("timestamp", datetime.utcnow().isoformat())
            dict_to_xml(data, root)
            
            # Pretty print XML
            rough_string = ET.tostring(root, encoding='unicode')
            reparsed = minidom.parseString(rough_string)
            content = reparsed.toprettyxml(indent="  ")
            
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
            return content, filename
        except Exception as e:
            logger.error(f"Error exporting to XML: {str(e)}")
            raise
    
    @staticmethod
    def export_to_pdf(data: Dict, title: str = "Export Report") -> Tuple[bytes, str]:
        """
        Export data as PDF
        
        Args:
            data: Data dictionary
            title: Report title
            
        Returns:
            Tuple of (content_bytes, filename)
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab not installed")
        
        try:
            buffer = BytesIO()
            pdf = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter
            
            # Title
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(inch, height - inch, title)
            
            # Timestamp
            pdf.setFont("Helvetica", 10)
            pdf.drawString(inch, height - 1.3*inch, f"Generated: {datetime.now().isoformat()}")
            
            # Content
            pdf.setFont("Helvetica", 10)
            y = height - 1.8 * inch
            
            for key, value in data.items():
                pdf.drawString(inch, y, f"{key}: {str(value)[:50]}")
                y -= 0.3*inch
                
                if y < inch:
                    pdf.showPage()
                    y = height - inch
            
            pdf.save()
            content = buffer.getvalue()
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return content, filename
        except Exception as e:
            logger.error(f"Error exporting to PDF: {str(e)}")
            raise
    
    @staticmethod
    def export_to_xlsx(data: Dict) -> Tuple[bytes, str]:
        """
        Export data as Excel file
        
        Args:
            data: Data dictionary or list of dicts
            
        Returns:
            Tuple of (content_bytes, filename)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed")
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Style for headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            if isinstance(data, dict):
                # Single record - write as key-value pairs
                ws.append(["Field", "Value"])
                ws[1][0].fill = header_fill
                ws[1][0].font = header_font
                ws[1][1].fill = header_fill
                ws[1][1].font = header_font
                
                for key, value in data.items():
                    ws.append([key, str(value)])
            
            elif isinstance(data, list) and data and isinstance(data[0], dict):
                # Multiple records - write as table
                headers = list(data[0].keys())
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row in data:
                    ws.append([str(row.get(h, "")) for h in headers])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
            
            buffer = BytesIO()
            wb.save(buffer)
            content = buffer.getvalue()
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return content, filename
        except Exception as e:
            logger.error(f"Error exporting to XLSX: {str(e)}")
            raise
    
    @staticmethod
    def export_to_sql(data: Dict, table_name: str = "data") -> Tuple[str, str]:
        """
        Export data as SQL INSERT statements
        
        Args:
            data: Data dictionary or list of dicts
            table_name: Table name for INSERT statements
            
        Returns:
            Tuple of (content, filename)
        """
        try:
            sql_statements = []
            
            if isinstance(data, dict):
                data = [data]  # Wrap single record
            
            if not data:
                return "", "export.sql"
            
            # Generate CREATE TABLE statement
            first_record = data[0]
            columns = list(first_record.keys())
            create_sql = f"CREATE TABLE IF NOT EXISTS {table_name} (\n"
            create_sql += "  id INTEGER PRIMARY KEY AUTOINCREMENT,\n"
            create_sql += ",\n".join([f"  {col} TEXT" for col in columns])
            create_sql += "\n);\n\n"
            sql_statements.append(create_sql)
            
            # Generate INSERT statements
            for record in data:
                cols = ",".join(columns)
                vals = ",".join([f"'{str(v).replace(chr(39), chr(39)*2)}'" for v in record.values()])
                insert_sql = f"INSERT INTO {table_name} ({cols}) VALUES ({vals});"
                sql_statements.append(insert_sql)
            
            content = "\n".join(sql_statements)
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
            return content, filename
        except Exception as e:
            logger.error(f"Error exporting to SQL: {str(e)}")
            raise
    
    @staticmethod
    def get_export_content(
        data: Dict,
        format: str,
        **kwargs
    ) -> Tuple[Any, str, str]:
        """
        Export data in specified format
        
        Args:
            data: Data to export
            format: Export format (json, csv, xml, pdf, xlsx, sql)
            **kwargs: Additional format-specific arguments
            
        Returns:
            Tuple of (content, filename, content_type)
        """
        format = format.lower()
        
        if format == "json":
            content, filename = ExportService.export_to_json(data)
            content_type = "application/json"
        
        elif format == "csv":
            content, filename = ExportService.export_to_csv(data, kwargs.get("headers"))
            content_type = "text/csv"
        
        elif format == "xml":
            content, filename = ExportService.export_to_xml(data, kwargs.get("root_name", "export"))
            content_type = "application/xml"
        
        elif format == "pdf":
            content, filename = ExportService.export_to_pdf(data, kwargs.get("title", "Export"))
            content_type = "application/pdf"
        
        elif format == "xlsx":
            content, filename = ExportService.export_to_xlsx(data)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        elif format == "sql":
            content, filename = ExportService.export_to_sql(data, kwargs.get("table_name", "data"))
            content_type = "application/sql"
        
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        return content, filename, content_type


# Singleton instance
_export_service = None


def get_export_service() -> ExportService:
    """Get or create export service instance"""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service